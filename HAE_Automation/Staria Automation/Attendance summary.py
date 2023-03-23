# -*-coding:utf-8 -*-
"""
# File     : Attendance_pay_calculation.py
# Time     ：1/3/2023 14:48
# Author   ：ZIYE Night
# version  ：python 3.10
# py文件说明：
    考勤系统 - 考勤记录汇总、计算
    2023-3-23: 调整统计，新增Actual work day
"""

import os
import re
from time import sleep
from itertools import chain
from datetime import date, timedelta
from tkinter import PhotoImage, Tk, Label, Button, Entry
from tkinter import GROOVE, CENTER, END
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox as tm

from openpyxl import load_workbook as load
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from warnings import filterwarnings

filterwarnings('ignore')


class Initialization:
    def __init__(self, _file_path, init_window_name, progress_bar):
        self._file_path = _file_path
        self.EmployeeName = []
        self.Header = ['Name',
                       'Days Present',
                       'Overtime leave(Hour)',
                       'Person leave(Day)',
                       'Sick leave(Day)']
        self.unique_dict = {}
        self._Max_row, self._Max_column = 1, 1
        self.outfit_list = []

        self.progress_bar = progress_bar
        self.init_window_name = init_window_name

    def _get_total_data(self):
        workbook_object = load(filename=self._file_path)
        sheet_names = workbook_object.sheetnames
        sheet_object = workbook_object[sheet_names[0]]
        self._Max_row, self._Max_column = sheet_object.max_row, sheet_object.max_column

        for row_index in range(5, sheet_object.max_row + 1):
            name = sheet_object.cell(row=row_index, column=1).value.rstrip()
            self.EmployeeName.append(name)

        for column_index in range(1, sheet_object.max_column + 1):
            outfit = sheet_object.cell(row=3, column=column_index).value
            if outfit:
                self.outfit_list.append(outfit.rstrip())
            else:
                self.outfit_list.append(sheet_object.cell(row=4, column=column_index).value.rstrip())

        _total_att_data = []
        for row in range(1, self._Max_row + 1):
            self.progress_bar['value'] = 50 / (self._Max_row + 1) * row
            self.init_window_name.update_idletasks()
            emp_name = sheet_object['A{}'.format(row)].value
            unique_id = sheet_object['D{}'.format(row)].value
            if unique_id and bool(re.search(r'\d', unique_id)):
                self.unique_dict[emp_name] = unique_id

            if emp_name and emp_name.rstrip() in self.EmployeeName:
                month_data = []
                for column in range(1, self._Max_column + 1):
                    cell_value = sheet_object.cell(row, column).value
                    if cell_value:
                        month_data.append(cell_value)
                    else:
                        month_data.append(0)
                _total_att_data.append(dict(zip(self.outfit_list, month_data)))
        _need_data = list(map(self._get_single_data, _total_att_data))

        return _need_data

    def _get_single_data(self, dict_data):
        single_dict, day_data = ({} for _ in range(2))

        def _set_dict_kv(sub_key, sub_value, sub_dict: dict) -> dict:
            if sub_key not in sub_dict.keys():
                sub_dict[sub_key] = [sub_value]
            else:
                sub_dict[sub_key].append(_)
            return sub_dict

        for k, v in dict_data.items():
            v = str(v).rstrip()
            if k in self.Header:
                single_dict[k] = v
            else:
                for _ in v.split('\n'):
                    if 'Clock In Late' in _:
                        _set_dict_kv(k, _, day_data)
                    elif 'Clock Out Leave Early' in _:
                        _set_dict_kv(k, _, day_data)
                    elif 'Off-site' in _:
                        _set_dict_kv(k, _, day_data)
                    elif 'Absenteeism' in _:
                        _set_dict_kv(k, _, day_data)
                    elif 'Clock In Absent' in _:
                        _set_dict_kv(k, _, day_data)
                    else:
                        if bool(re.search(r'\d', k)):
                            day_data.setdefault(k, [])

        _absence_time_total = map(self._get_absence_time, day_data.values())
        new_values = map(lambda x: x[0] + x[1], zip(list(day_data.values()), _absence_time_total))
        day_data = dict(zip(day_data.keys(), new_values))

        new_single_list = list(single_dict.items())
        # new_single_list.insert(2, ('Actual work day', '0'))
        merge_dict = dict(new_single_list + list(day_data.items()))
        return merge_dict

    def _get_absence_time(self, day_values):

        def _time_conversion(time_value):
            res_time = []
            if 'Hours' in time_value and 'Mins' in time_value:
                total_time = [int(_.rstrip()) for _ in re.findall(r'(.*)Hours(.*)Mins', time_value)[0]]
                res_time = sum([total_time[0] * 60, total_time[1]])
            elif 'Hours' in time_value:
                total_time = [int(_.rstrip()) for _ in re.findall(r'(.*)Hours', time_value)[0]]
                res_time = sum([total_time[0] * 60])
            elif 'Mins' in time_value:
                total_time = [int(_.rstrip()) for _ in re.findall(r'(.*)Mins', time_value)]
                res_time = sum(total_time)
            return res_time

        if not day_values:
            return [str(0)]
        else:
            _origin_time = []
            for value in day_values:
                if 'Late' in value or 'Leave Early' in value:
                    late_time, leave_early_time = re.findall(r"Late(.*)", value), re.findall(r"Leave Early(.*)", value)
                    _origin_time.append(late_time)
                    _origin_time.append(leave_early_time)
            _origin_time = list(chain(*_origin_time))
            _absence_time = sum(list(map(_time_conversion, _origin_time)))
            return [str(0)] if not _absence_time else [str(_absence_time)]


class Salary:
    def __init__(self, year, month):
        self.year = year
        self.month = month
        self.work_days = self.Daily_Wage()

    # 计算工作天数
    def Daily_Wage(self):
        # 获取该月份的第一天
        first_day = date(self.year, self.month, 1)
        # 获取该月份的最后一天 每月28号 + 4天。
        last_day = date(self.year, self.month, 28) + timedelta(days=4)
        # 将最后一天设置为该月份的最后一天
        last_day = last_day - timedelta(days=last_day.day)

        # 统计周六和周日的数量
        num_fri, num_sun = 0, 0
        delta = timedelta(days=1)
        curr_day = first_day
        while curr_day <= last_day:
            if curr_day.weekday() == 4:  # 4表示周五
                num_fri += 0.5
            elif curr_day.weekday() == 6:  # 6表示周日
                num_sun += 1
            curr_day += delta

        holiday_days = num_sun + num_fri
        work_days = float(str(last_day).split("-")[-1]) - holiday_days
        return work_days

    """
    计算迟到早退扣款
    """
    def Late(self, lates):
        if len(lates) <= 2: return 0
        lateday = list(lates.keys())
        latetime = list(lates.values())
        # 统计外勤打卡：
        Clock_In_Offsite = 0
        pattern_str1 = re.compile(r'\bClock In Off-site\b')
        pattern_str2 = re.compile(r'\b"Clock Out Off-site"\b')
        # 缺勤统计
        Clock_dict = {}
        late_str1 = re.compile(r'\bAbsence time\b')
        # 漏打卡统计
        Clock_Out_Leave = 0
        Leave_str1 = re.compile(r'\bClock In Absent\b')
        # 旷工统计
        Absenteeism = 0
        Absenteeism_str = re.compile(r'\bAbsenteeism\b')

        # 迟到总时长
        num, late_time = 0, 0
        while num < len(latetime):
            str_Data = " ".join(latetime[num])
            if re.search(pattern_str1, str_Data) or re.search(pattern_str2, str_Data):
                Clock_In_Offsite += 1
            if re.search(Absenteeism_str, str_Data):
                Absenteeism += 1
            if re.search(Leave_str1, str_Data):
                Clock_Out_Leave += 0.5

            Min = float(latetime[num][-1])
            late_time += Min
            if 0 < Min < 240:
                Clock_dict[lateday[num]] = 0.5
            elif Min > 240:
                Clock_dict[lateday[num]] = 1
            num += 1

        # 迟到天数切片
        days_list, time_list = [], []
        Clock_list_keys = list(Clock_dict.keys())
        Clock_list_values = list(Clock_dict.values())

        for day_ in range(len(Clock_list_keys)):
            # days = Clock_list_keys[day_]
            late_nums = Clock_list_values[day_]
            if not time_list:
                time_list.append([late_nums])
            elif float(Clock_list_keys[day_][3:]) == float(Clock_list_keys[day_ - 1][3:]) + 1:
                time_list[-1].append(late_nums)
            else:
                time_list.append([late_nums])

        New_Clock_In_Late = 0
        for nu in time_list:
            if len(nu) >= 3:
                New_Clock_In_Late += len(nu) * 1
            else:
                New_Clock_In_Late += sum(nu)
        Violation_Days = float((New_Clock_In_Late + Clock_Out_Leave + Absenteeism))
        # 返回值-外勤打卡天数，迟到算换天数，迟到总时长，漏打卡天数， 旷工天数, 迟到扣除工资
        Dicts_keys = ['Clock In Offsite(Day)', 'Clock In Late(Day)', 'Clock In Late Time(Mins)', 'Clock Out Leave(Day)', 'Absenteeism(Days)', 'Violation Days']
        Dicts_values = [Clock_In_Offsite, New_Clock_In_Late, late_time, Clock_Out_Leave, Absenteeism, Violation_Days]
        Lates_Dict = dict(zip(Dicts_keys, Dicts_values))
        return Lates_Dict

    '''
    Main方法
    '''

    def Dict_main(self, sum_dict):
        Keys = [_data for _data in sum_dict.keys()]
        # print('\nKeys: ', Keys)
        Values = [_data for _data in sum_dict.values()]
        # print('\nValues: ', Values)
        Actual_values = Values[1]
        Dict_late_data = {}  # 迟到字典
        Dict_Leave_day = Values[3]  # 请假
        for i in range(len(Keys[5:-1])):
            Dict_late_data[Keys[5:-1][i]] = Values[5:-1][i]

        late_salary = self.Late(Dict_late_data)  # 违规信息
        late_salary_key = list(late_salary.keys())
        late_salary_value = list(late_salary.values())

        conutkey, countvalues = Keys[0:5], Values[0:5]
        conutkey.extend(late_salary_key)
        countvalues.extend(late_salary_value)

        conutkey.insert(2, "Actual work day")
        countvalues.insert(2, Actual_values)
        CountDict = dict(zip(conutkey, countvalues))
        CountDict['Violation Days'] = late_salary_value[-1] + float(Dict_Leave_day)
        CountDict['Days Present'] = self.work_days
        return CountDict


class WriteExcel:
    def __init__(self, origin_data, month, init_window_name, progress_bar):
        self._write_data = self._merge_dict(origin_data, {})
        self.month = month

        self.progress_bar = progress_bar
        self.init_window_name = init_window_name

    def _merge_dict(self, origin_dict, result_dict):
        dict_index = 0
        _m1 = origin_dict[dict_index]
        for k, v in _m1.items():
            if k not in result_dict.keys():
                result_dict[k] = [v]
            else:
                result_dict[k].append(v)
        origin_dict.pop(dict_index)
        if len(origin_dict):
            return self._merge_dict(origin_dict, result_dict)
        else:
            return result_dict

    def _write_excel(self):
        cell_head_values, cell_values = list(self._write_data.keys()), list(self._write_data.values())
        alphabet = dict(zip(range(1, len(cell_head_values) + 1), [chr(letters) for letters in range(65, 91)]))
        filename = f'Attendance Summary - {self.month}.xlsx'
        wb_write = Workbook() if not os.path.exists(filename) else load(filename)
        sheet_names = wb_write.sheetnames
        sheet_wr_obj = wb_write.create_sheet(title='Monthly Summary', index=0) if 'Monthly Summary' not in sheet_names else wb_write['Monthly Summary']
        list_of_title_index = range(len(cell_head_values))
        title_start, title_end = list_of_title_index.start + 1, list_of_title_index.stop
        top_cell = sheet_wr_obj.cell(row=title_start, column=1, value='Summary of Attendance in February')

        for column_index in range(1, len(cell_head_values) + 1):
            sheet_wr_obj.column_dimensions[alphabet[column_index]].width = 26
        sheet_wr_obj.row_dimensions[title_start].height = 50

        sheet_wr_obj.merge_cells(start_row=1, start_column=title_start, end_row=1, end_column=title_end)
        top_cell.alignment = Alignment(horizontal='center', vertical='center')
        top_cell.fill = PatternFill("solid", fgColor='c3b091')
        top_cell.font = Font(name="微软雅黑", size=30, color="000000", bold=True, italic=True, strike=None, underline=None)

        head_index = 2
        sheet_wr_obj.row_dimensions[head_index].height = 25
        for header_index, header in enumerate(cell_head_values):
            single_head_cell = sheet_wr_obj.cell(row=head_index, column=header_index + 1, value=header)
            single_head_cell.alignment = Alignment(horizontal='center', vertical='center')
            single_head_cell.fill = PatternFill("solid", fgColor='ffc0cb')
            single_head_cell.font = Font(name="Malgun Gothic", size=12, color="87756e", bold=True, italic=None, strike=None, underline=None)

        for _index in range(len(cell_values)):
            self.progress_bar['value'] = 55 + (50 / len(cell_values) * _index)
            self.init_window_name.update_idletasks()
            for sub_index, sub_value in enumerate(cell_values[_index]):
                if type(sub_value) is str:
                    sub_value = float(sub_value) if sub_value.isdigit() or '.' in sub_value else sub_value
                single_cell = sheet_wr_obj.cell(row=head_index + sub_index + 1, column=_index + 1, value=sub_value)

                if _index + 1 == len(cell_values):
                    single_cell.fill = PatternFill("solid", fgColor='de5285')
        path = str(os.path.abspath(os.curdir)).split("\\")[:-1]
        path.append("Attendance summary sheet")
        path_name = "\\".join(path)
        if not os.path.exists(path_name):
            os.mkdir(path_name)
        wb_write.save(f'{path_name}\\{filename}')
        wb_write.close()


class ClockGUI:
    def __init__(self, init_window_name):
        self.init_window_name = init_window_name
        self.file_path = None
        self.Monthly_table = {'01': [1, 'January'], '02': [2, 'February'], '03': [3, 'March'],
                              '04': [4, 'April'], '05': [5, 'May'], '06': [6, 'June'],
                              '07': [7, 'June'], '08': [8, 'August'], '09': [9, 'September'],
                              '10': [10, 'October'], '11': [11, 'November'], '12': [12, 'December']}

    def TKset_init(self):
        self.init_window_name.title("考勤核算系统 V0.28")  # 窗口名
        self.init_window_name.iconbitmap("Img/Clock.ico")  # 设置窗口图标
        self.init_window_name.geometry('360x210+50+40')  # 窗口大小以及定位点
        self.init_window_name.attributes('-topmost', True)  # 窗口置顶
        self.init_window_name.resizable(width=False, height=False)

        # 背景
        global pil_image1
        pil_image1 = PhotoImage(file="Img/Clock.png")
        self.init_window_img = Label(self.init_window_name, image=pil_image1)
        self.init_window_img.place(x=-2, y=5)  # x：定位坐标； relx：百分比定位坐标；relwidth： 长宽
        # 公司logo
        global pil_image2
        pil_image2 = PhotoImage(file="Img/HAE 图标.png")
        self.init_window_img2 = Label(self.init_window_name, image=pil_image2)
        self.init_window_img2.place(x=300, y=15)

        # 标签
        self.init_data_label_A = Label(self.init_window_name, text="Please select the \nattendance file: ", bg="white",
                                       height=2, width=18, font=("黑体", 12), anchor='e', justify='center')
        self.init_data_label_A.grid(row=0, column=0, padx=2, pady=10, ipady=2)
        self.init_data_label_filename = Label(self.init_window_name, text="Attendance Sheet: ", bg="whitesmoke",
                                              height=1, width=18, font=("黑体", 11), anchor='e', justify='center')
        self.init_data_label_filename.grid(row=1, column=0, padx=2, pady=2, ipady=2, sticky='w')
        self.init_data_label_explain = Label(self.init_window_name, text="注：欢迎使用，有任何问题请联系HAE-PY团队", bg="whitesmoke", font=("黑体", 8), anchor='e', justify='center')
        self.init_data_label_explain.place(x=0, y=182, height=28, width=360)

        # 间隔标签
        self.lb_blue = Label(self.init_window_name, text="--------", bg="blue", fg='#ffffff', relief=GROOVE)
        self.lb_blue.place(x=0, y=100, height=5, width=360)

        # 按钮
        self.init_data_botton1 = Button(self.init_window_name, text="select files", height=1, width=13, bd=6, bg="honeydew", justify=CENTER, command=self.choose_file)
        self.init_data_botton1.grid(row=0, column=1, padx=0, pady=5, ipady=2)
        self.init_data_botton1.place(x=160, y=14)
        self.init_data_botton_Operation = Button(self.init_window_name, text="Operation", height=1, width=10, bd=4, bg="gray", activebackground='teal', justify=CENTER, command=self.main)
        self.init_data_botton_Operation.place(x=80, y=120)
        self.init_data_botton_Quit = Button(self.init_window_name, text="Quit", height=1, width=10, bd=4, bg="gray", justify=CENTER, command=self.init_window_name.quit)
        self.init_data_botton_Quit.place(x=200, y=120)

        # 文本框
        self.init_data_entry_fliename = Entry(self.init_window_name, bd=3, width=27)
        self.init_data_entry_fliename.grid(row=1, column=1, padx=5, ipady=2, sticky='nw')

        # 进度条
        self.progress_bar = ttk.Progressbar(self.init_window_name, orient="horizontal", length=200, mode="determinate")
        self.progress_bar.place(x=20, y=160, height=20, width=325)

    def choose_file(self):
        # 打开文件选择器窗口
        self.file_path = filedialog.askopenfilename(title="选择文件", filetypes=(("Excel文件", "*.xlsx;*.xls"),))
        if self.file_path:
            self.init_data_entry_fliename.delete(0, END)
            self.init_data_entry_fliename.insert(0, self.file_path.split("/")[-1])
        else:
            tm.showwarning('Warning', 'No folder selected, please select again!')

    def main(self):
        Init = Initialization(self.file_path, self.init_window_name, self.progress_bar)
        try:
            if self.file_path:
                date_data = re.findall(r"\d+", self.file_path)[-1]
                year = int(date_data[0:4])
                month = self.Monthly_table[date_data[4:][0:2]][0]
                format_month = self.Monthly_table[date_data[4:][0:2]][1]
                combined_data = list(map(Salary(year, month).Dict_main, Init._get_total_data()))

                WriteExcel(combined_data, format_month, self.init_window_name, self.progress_bar)._write_excel()
                sleep(0.25)
                tm.showinfo(title='Tip!', message='Successfully generated {} Attendance Sheet'.format(format_month))
            else:
                tm.showerror(title='Error!', message='No files have been selected')
        except:
            tm.showerror(title='Error!', message='Program failed to run, please contact the development team')


init_window = Tk()
Windows = ClockGUI(init_window)
# 设置根窗口默认属性
Windows.TKset_init()
init_window.mainloop()
