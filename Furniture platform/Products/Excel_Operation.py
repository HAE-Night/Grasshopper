from typing import List

import os
import shutil
import copy
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, Border, Side
from Products import models



#创建一个List_Products类是List类型的Products

class ExcelCopier:
    def __init__(self, excel_path, sheet_title='Sheet1'):
        self.wb = openpyxl.load_workbook(excel_path, data_only=True)
        self.sheet_title = sheet_title

    def Write_To_Excel(self, List_Products):
        """
        将数据写入到excel表格中
        :param headers: 表头
        :param data: 表格信息
        :return:
        """
        #循环遍历List_Products的值和下标
        for Product_index, Product in enumerate(List_Products):
            product_id = Product[0]
            product_num = Product[1]
            product_discount = Product[2]
            product_data = models.Products.objects.get(ID=product_id)
            product_item = product_data.ITEM
            product_photo = product_data.ITEM_PHOTO
            product_room_area = product_data.ROOM_AREA
            product_description = product_data.DESCRIPTION
            product_size_dimension = product_data.SIZE_DIMENSION
            product_length_mm = product_data.LENGTH_MM
            product_width_mm = product_data.WIDTH_MM
            product_height_mm = product_data.HEIGHT_MM
            product_unit = product_data.UNIT
            product_unit_price_aed = product_data.UNIT_PRICE_AED
        # # 以下是写入图片和计算总价
        # listDoc = os.listdir(self.save_path)
        # listDoc.sort(key=lambda x: int(x.split('.')[0]))
        # for i in range(len(listDoc)):
        #     listDoc[i] = self.save_path + listDoc[i]
        # for _i in range(len(listDoc)):
        #     img = Image(listDoc[_i])
        #     img_zoom = 120 / img.width
        #     img.width, img.height = 120, img.height * img_zoom
        #     img.anchor = 'center'
        #     cell = ws.cell(row=_i + 8, column=Photo_index)
        #     ws.row_dimensions[_i + 8].height = img.height / 18 * 13.5 + 5
        #     ws.add_image(img, cell.coordinate)
        #     Quantity_cell = ws.cell(_i + 8, Quantity_index)
        #     Price_cell = ws.cell(_i + 8, Price_index)
        #     Total_cell = ws.cell(_i + 8, len(headers) + 1)
        #     try:
        #         Quantity_cell.value = int(Quantity_cell.value)
        #         Price_cell.value = float(Price_cell.value)
        #         Total_cell.value = '={}*{}'.format(Quantity_cell.coordinate, Price_cell.coordinate)  # 写入公式
        #     except Exception:
        #         Total_cell.value = 'N/A'
        #     Total_cell.font = font
        #     Total_cell.alignment = align
        #
        # Total_Amount_cell = ws.cell(row=ws.max_row + 1, column=ws.max_column)  # 设置总价的单元格
        # Total_Amount_cell.font = font
        # Total_Amount_cell.alignment = align
        # Total_Amount_cell.value = '=SUM({}:{})'.format(ws.cell(row=8, column=ws.max_column).coordinate,
        #                                                ws.cell(row=ws.max_row - 1, column=ws.max_column).coordinate)
        #
        # # # 以下是设置边框（必须要设置了框线才能合并单元格）
        # for row in ws.iter_rows(min_row=ws.max_row):
        #     for cell in row:
        #         # 根据需求选择性地应用边框样式
        #         cell.border = border
        #
        # # 以下是合并表格
        # ws.merge_cells(start_row=ws.max_row, start_column=1,
        #                end_row=ws.max_row, end_column=ws.max_column - 1)
        # Merge_cell = ws.cell(row=ws.max_row, column=1)
        # Merge_cell.font = Font(name='等线', size=9, bold=True)
        # Merge_cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        # Merge_cell.value = 'Total Amount'
        #
        # self.Add_Note(ws)
        # ws.page_setup.fitToPage = True
        # ws.page_setup.fitToHeight = 0
        # ws.page_setup.fitToWidth = 1
        #
        # # 以下是保存excel表格
        # # new_path = r'D:\Work\平台\家具自动化\案例\Test.xlsx'
        # self.wb.save(excel_path)
        # self.wb.close()
        # shutil.rmtree(self.save_path)  # 删除临时文件夹




# # 使用上述三个类来实现整个程序
# reader = ExcelReader('source.xlsx')
# copier = ExcelCopier(reader, 'destination.xlsx')
# copier.copy_data()
# pdf_converter = ExcelToPdfConverter('destination.xlsx')
# pdf_converter.convert()
if __name__ == '__main__':
    # ex_pdf = ExcelToPdfConverter(r"E:\Desktop\Test.xlsx")
    # ex_pdf.convert(r"E:\Desktop\Test.pdf", "Sheet")
    path = r"E:\Documents\家具报价\报价分析案例\506.152\4.COST CHECKING-INTERNAL\BOQ 152 from Cost checking.xlsx"
    Quotation_NO = 'HAE'
    Note = ['1. 请注意，本表格中所列的价格为参考价，请以实际成交价为准。',
            '2. 如有疑问，请联系客户经理。',
            '3. 如有需要']
    new_sheet_name = 'sheet221'
    sheet_bom = "Cost & DXB Market Price 12.19"
    Excel = ExcelCopier(excel_path=path, Quotation_NO='HAE', Note=Note, sheet_bom=sheet_bom, sheet_title=new_sheet_name)
    Excel.Download_Excel_Image(r"E:\Pictures\new\\")
    headers, data = Excel.Get_Excel_Info()
    Excel.Write_To_Excel(headers, data, path)  # 将数据写入到excel表格中
