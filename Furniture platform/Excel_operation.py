import os
from itertools import zip_longest

import openpyxl
from openpyxl.utils import get_column_letter
from PIL import Image as pil_image


class ExcelReader:
    def __init__(self, filename):
        self.wb = openpyxl.load_workbook(filename)

    def get_sheet_names(self):
        return self.wb.sheetnames

    def get_image_path(self, cell):
        pass
        # return get_image_path(cell.value)

    def get_data(self, cell):
        return cell.value

    # 自动关闭excel文件
    def close(self):
        self.wb.close()


class ExcelCopier:
    def __init__(self, excel_path, sheet_bom):
        self.wb = openpyxl.load_workbook(excel_path, data_only=True)
        self.save_path = None
        self.sheet_bom = sheet_bom

    def Download_Excel_Image(self, img_path, start_index=0):
        wb = self.wb
        PIL_Image_List = []
        Image_info_row_list = []
        img_name_list = []
        ws = wb[self.sheet_bom]  # 获取第一个sheet

        save_path = img_path  # 保存路径
        if not os.path.exists(save_path):
            os.mkdir(save_path)  # 创建文件夹

        for i, image in enumerate(ws._images):
            img = pil_image.open(image.ref).convert("RGB")  # 打开图片
            PIL_Image_List.append(img)
            image_info = image.anchor._from  # 图片的位置
            col, row = (get_column_letter(image_info.col + 1), image_info.row)
            img_name_list.append(str(row + start_index))
            Image_info_row_list.append(row)
        sorted_indices = [i for i, _ in sorted(enumerate(Image_info_row_list), key=lambda x: x[1])]  # 将图片数据根据行数进行排序
        combined_lists = [PIL_Image_List[i] for i in sorted_indices]
        img_name_list_1 = [img_name_list[i] for i in sorted_indices]

        list_save_path = [(save_path + img_name_list_1[_] + '.png', _1) for _, _1 in enumerate(combined_lists)]

        # 将图片保存到指定位置
        for img_path, img in list_save_path:
            print(img_path)
            img.save(img_path)
        return list(set(img_name_list_1))

    def Get_Excel_Info(self):
        """
        读取Excel表格中的信息
        :param sheet_name:
        :return: headers: 表头
                result: 表格信息
        """
        header_row = 1  # 设置表头在第几行
        ws = self.wb[self.sheet_bom]  # 获取第一个sheet
        Mate_CATEGORY = ["CATEGORY", "CATALOGUE"]
        Mate_ROOM_AREA = ["AREA", "ROOM/AREA", "ROOMAREA"]
        Mate_Item = ["ITEMNAME", "ITEMSNAME", "ITEM", "PRODUCTDESCRIPTION"]  # 商品名称列表
        Mate_Description = ["DESCRIPTION", "REFERENCEDDATAANDSPECIFICATIONS", "REFERENCEDDATAANDSPECIFICATION",
                            "SPECIFICATIONS", "REFERENCEDDESCRIPTION", "FABRIC"]  # 商品描述列表
        Mate_SIZE_DIEMENSION = ["SIZE", "APPROX.DIMENSIONS", "DIMENSIONS", "SPECIFICATION"]
        Mate_Length = ["LENGTH", "LENGTH(MM)"]  # 长度列表
        Mate_Width = ["WIDTH", "WIDTH(MM)", "DEPTH(MM)"]  # 宽度列表
        Mate_Height = ["HEIGHT", "HEIGHT(MM)"]  # 高度列表
        Mate_Quantity = ["QUANTITY(QTY)", "QTY", "QUANTITY"]  # 数量列表
        Mate_Remark = ["REMARK(UNIT)", "UNITS", "UNIT"]  # 备注列表
        Mate_Unit_price = ["UNITPRICE", "UNITPRICE(AED)", "UNITPRICE/AED", "UNITPRICEAED"]  # 单价列表
        # 将Item转换为大写
        List_Category = []
        List_Room_Area = []
        List_Item = []  # 商品名称列表
        List_Description = []  # 商品描述列表
        List_Size_Dimension = []
        List_Length = []  # 长度列表
        List_Width = []  # 宽度列表
        List_Height = []  # 高度列表
        List_Quantity = []  # 数量列表
        List_Remark = []  # 备注列表
        List_Unit_price = []  # 单价列表
        #  获取到在Excel表格中的所有表头，并得到其字母编号
        for cell in ws[header_row]:
            cell_value = str(cell.value).replace(" ", "").replace("\n", "").upper()
            header_number = cell.coordinate

            if cell_value in Mate_CATEGORY:  # 如果商品名称在列标题中存在
                for col in ws.iter_cols(min_col=ws[header_number].column, min_row=header_row + 1,
                                        max_col=ws[header_number].column):
                    for cell_ in col:
                        List_Category.append(cell_.value)

            if cell_value in Mate_ROOM_AREA:  # 如果商品名称在列标题中存在
                for col in ws.iter_cols(min_col=ws[header_number].column, min_row=header_row + 1,
                                        max_col=ws[header_number].column):
                    for cell_ in col:
                        List_Room_Area.append(cell_.value)

            if cell_value in Mate_Item:  # 如果商品名称在列标题中存在
                for col in ws.iter_cols(min_col=ws[header_number].column, min_row=header_row + 1,
                                        max_col=ws[header_number].column):
                    for cell_ in col:
                        List_Item.append(cell_.value)

            if cell_value in Mate_Description:  # 如果商品名称在列标题中存在
                for col in ws.iter_cols(min_col=ws[header_number].column, min_row=header_row + 1,
                                        max_col=ws[header_number].column):
                    for cell_ in col:
                        _ = cell_.value
                        if _:
                            List_Description.append(_.replace("\n", ""))
                        else:
                            List_Description.append(_)

            if cell_value in Mate_SIZE_DIEMENSION:  # 如果商品名称在列标题中存在
                for col in ws.iter_cols(min_col=ws[header_number].column, min_row=header_row + 1,
                                        max_col=ws[header_number].column):
                    for cell_ in col:
                        List_Size_Dimension.append(cell_.value)

            if cell_value in Mate_Length:  # 如果商品名称在列标题中存在
                for col in ws.iter_cols(min_col=ws[header_number].column, min_row=header_row + 1,
                                        max_col=ws[header_number].column):
                    for cell_ in col:
                        List_Length.append(cell_.value)

            if cell_value in Mate_Width:  # 如果商品名称在列标题中存在
                for col in ws.iter_cols(min_col=ws[header_number].column, min_row=header_row + 1,
                                        max_col=ws[header_number].column):
                    for cell_ in col:
                        List_Width.append(cell_.value)

            if cell_value in Mate_Height:  # 如果商品名称在列标题中存在
                for col in ws.iter_cols(min_col=ws[header_number].column, min_row=header_row + 1,
                                        max_col=ws[header_number].column):
                    for cell_ in col:
                        List_Height.append(cell_.value)

            if cell_value in Mate_Quantity:  # 如果商品名称在列标题中存在
                for col in ws.iter_cols(min_col=ws[header_number].column, min_row=header_row + 1,
                                        max_col=ws[header_number].column):
                    for cell_ in col:
                        List_Quantity.append(cell_.value)

            if cell_value in Mate_Remark:  # 如果商品名称在列标题中存在
                for col in ws.iter_cols(min_col=ws[header_number].column, min_row=header_row + 1,
                                        max_col=ws[header_number].column):
                    for cell_ in col:
                        List_Remark.append(cell_.value)

            if cell_value in Mate_Unit_price:  # 如果商品名称在列标题中存在
                for col in ws.iter_cols(min_col=ws[header_number].column, min_row=header_row + 1,
                                        max_col=ws[header_number].column):
                    for cell_ in col:
                        List_Unit_price.append(cell_.value)

        List_result = list(
            zip_longest(List_Category, List_Room_Area, List_Item, List_Description, List_Size_Dimension, List_Length,
                        List_Width,
                        List_Height, List_Quantity, List_Remark, List_Unit_price,
                        fillvalue=None))  # 将各个列表进行打包，若长度不一致则用None填充

        return List_result


if __name__ == '__main__':
    path = r"E:\Documents\家具报价\报价分析案例\506.151\4.COST CHECKING-INTERNAL\B.O.Q_151 Cost checking.xlsx"
    sheet_bom = "151 BOQ from Egypt team"
    Excel = ExcelCopier(excel_path=path, sheet_bom=sheet_bom)
    # print(Excel.Download_Excel_Image(img_path="IMG/", start_index=0))
    Excel.Get_Excel_Info()
